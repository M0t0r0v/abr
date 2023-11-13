import sqlite3
import openpyxl as oxl
from re import findall
from tqdm import tqdm
import itertools

DATABASE: str = 'words.db'  # Имя Базы данных
TABLE: str = 'db_words'     # Имя таблицы данных

db = sqlite3.connect(DATABASE)  # Подсоединяем базу данных словарей
sql = db.cursor()


def search_word_in_db(word: str) -> str:
    """Функция перевода слов по базе данных DATABASE.
    Пробегает по всей таблице TABLE, ищем слова из field[1]-field[7].
    И берет найденую аббревиатуру из field[0]."""

    if word is not None:  # Если слово не пустое
        # sql запрос, все слова
        for field in sql.execute('SELECT * FROM ' + (TABLE)):
            if any(f == word for f in field[1:7]):
                # Подставляемая аббревиатура из колонки field1 словаря
                return field[0]


def is_all_upper(text: str) -> bool:
    """Проверка предложения на все маркировки и абревиатуры"""

    if text.upper() == text:
        return True
    elif len(text) == 0:
        return True
    return False


def spec_symb(text: str) -> bool:
    """Патерн поиска аббревиатур и спец последовательностей."""

    mark = r'(\b[а-я]{1}[!-~]{1}[а-я]{1}\b)'
    r'|(\b[А-Я]{2}[а-я]{1}\b)'  # например в тексте МПа КПа
    r'|(\b[0-9]{1}[а-я]{2}\b)'  # например в тексте 9я
    r'|(\b[A-Z]{1}[a-z]{1}[0-9]{1}\b)'
    r'|(\b[a-z!-~0-9]{6})'
    r'|(\b[A-Za-z]\b)'

    mark: list = findall(mark, text)

    if not mark:
        return False
    elif len(mark) > 0:
        return True
    return False


def translit(text) -> str:
    # спецфильтр
    replacements = [['*', ''],
                    ['#', ''],
                    ['“', ''],
                    ['”', ''],
                    ['–', ''],
                    ['%', ''],
                    ['~220В', '220VAC'],
                    ['=220В', '220VDC'],
                    ['=24В', '24VDC'],
                    ['°C', '']
                    ]
    # пробегаемся по тексту фильтром
    for frm, to in replacements:
        text = text.replace(frm, to)

    cyrillic = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя-/№><.'

    latin = 'a|b|v|g|d|e|e|z|z|i|i|k|l|m|n|o|p|r|s|t|u|f|x|tc|ch|sh|shch||y||e|iu|ia|_|_|N|more|less|_|'.split('|')  # таблица транслитерации
    trantab = {k: v for k, v in zip(cyrillic, latin)}
    newtext = ''
    for ch in text:
        casefunc = str.capitalize if ch.isupper() else str.lower
        newtext += casefunc(trantab.get(ch.lower(), ch))
    return str(newtext)


def main(s: str) -> str:
    """Функция отделения аббревиатур от основного предложения."""

    if s is not None:
        # фильтр опечаток
        replacement2 = [['. ', ' '],
                        ['-', ' '],
                        [',', ' '],
                        ['"', ' '],
                        ['(', ' '],
                        [')', ' '],
                        ['“', ''],
                        ['”', ''],
                        ['«', ''],
                        [';', ''],
                        ['»', ''],
                        ['  ', ' ']
                        ]
        for frm, to in replacement2:
            s = s.replace(frm, to)

        s = s.split()  # формируем исходный список слов из предложения

        numword: int = len(s)  # определяе колличество слов в предложении

        # формируем пустой список с числом элементов количества слов
        list1 = [''] * numword
        list2 = [''] * numword

        for i in range(0, numword):  # записываем аббревиатуры в пустой список
            if is_all_upper(str(s[i])) or spec_symb(str(s[i])):
                list1[i] = str(s[i])
                list2[i] = translit(str(s[i]))
        # исключаем из исходного списка аббревиатуры
        result = list(itertools.filterfalse(list1.__contains__, iter(s)))
        k = [''] * numword
        # формируем пустой список с числом элементов количества слов
        list3 = [''] * numword
        for g in range(0, len(result)):
            list3[g] = search_word_in_db(str(result[g].lower()))
            if list3[g] is None:
                list3[g] = ''
                k = result[g].lower()
                continue

        # Поэлементное слияние листа маркировок и аббревиатур
        list_c = []

        for i in range(numword):
            if list2[i] == '' and list3[i] != '':
                list_c.append(list3[i])
            elif list3[i] != '' and list2[i] != '':
                list_c.insert(i, list2[i])
                list_c.insert(i + 1, list3[i])
            elif list3[i] == '' and list2[i] != '':
                list_c.append(list2[i])
            elif list3[i] == '' and list2[i] == '':
                continue
        list_c = list(filter(lambda x: x != '', list_c))
        # избавляемся от дубликатов
        list_c = list(dict.fromkeys(list_c))
        x = '_'.join([str(x) for x in list_c])
        return str(x), str(k)
    else:  # лучше явно чем не явно
        # Если входное слово None то заполняем пустой строкой
        empty1: str = ''
        empty2: str = ''
        return empty1, empty2


# Файл поиска и записи результата
filename_excel = ('LIST.xlsx')
wb2 = oxl.reader.excel.load_workbook(filename=filename_excel, data_only=True)
wb2.active = 0
sheet = wb2.active
# for i in range (1,sheet.max_row+1):
for i in tqdm(range(1, sheet.max_row+1)):
    c, g = main((sheet['A'+str(i)].value))
    sheet['B' + str(i)] = c  # готовая аббревиатура
    sheet['C' + str(i)] = g  # слова которые на нашлись в словаре
wb2.save(filename_excel)
